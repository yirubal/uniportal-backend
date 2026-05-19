from datetime import timedelta
from io import BytesIO, StringIO
import tempfile

from django.contrib import admin as django_admin
from django.core.management import call_command, CommandError
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APITestCase

from apps.accounts.models import Student
from apps.api.views import _generate_jwt
from apps.content.models import Course, Department
from apps.quiz.admin import ExamPaperAdmin
from apps.quiz.engine import get_topics_with_low_score
from apps.quiz.importers import import_exit_questions_from_excel, import_questions_from_excel
from apps.quiz.models import Chapter, ExamPaper, Question, QuizAttempt


class QuizFeedbackApiTests(APITestCase):
    def setUp(self):
        self.student = Student.objects.create(
            telegram_id=123456789,
            first_name='Quiz',
            username='quizuser',
        )
        self.other_student = Student.objects.create(
            telegram_id=987654321,
            first_name='Other',
            username='otherquizuser',
        )
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {_generate_jwt(self.student)}')

        self.course = Course.objects.create(
            name='Programming Fundamentals',
            code='CS101',
        )
        self.exam_paper = ExamPaper.objects.create(
            title='Programming Quiz',
            course=self.course,
            exam_type=ExamPaper.TYPE_QUIZ,
            year=2026,
            access_level=ExamPaper.ACCESS_FREE,
        )
        self.first_question = Question.objects.create(
            exam_paper=self.exam_paper,
            text='What is 2 + 2?',
            option_a='3',
            option_b='4',
            option_c='5',
            option_d='6',
            correct_option='b',
            explanation='2 + 2 equals 4.',
            topic_tags=['arithmetic'],
        )
        self.second_question = Question.objects.create(
            exam_paper=self.exam_paper,
            text='Which keyword defines a function in Python?',
            option_a='func',
            option_b='define',
            option_c='def',
            option_d='lambda',
            correct_option='c',
            explanation='Python functions are declared with def.',
            topic_tags=['functions'],
        )

    def test_quiz_submission_stores_and_returns_detailed_answers(self):
        response = self.client.post(
            '/api/quiz/attempts/',
            {
                'course_id': self.course.id,
                'exam_paper_id': self.exam_paper.id,
                'mode': 'practice',
                'answers': [
                    {
                        'question_id': self.first_question.id,
                        'selected_option': 'b',
                    },
                    {
                        'question_id': self.second_question.id,
                        'selected_option': 'a',
                    },
                ],
            },
            format='json',
        )

        self.assertEqual(response.status_code, 201)
        attempt = QuizAttempt.objects.get(id=response.data['attempt_id'])
        detailed_answers = response.data['detailed_answers']
        first_key = str(self.first_question.id)
        second_key = str(self.second_question.id)

        self.assertEqual(attempt.detailed_answers, detailed_answers)
        self.assertEqual(detailed_answers[first_key]['selected_option'], 'b')
        self.assertEqual(detailed_answers[first_key]['correct_option'], 'b')
        self.assertTrue(detailed_answers[first_key]['is_correct'])
        self.assertEqual(detailed_answers[first_key]['question_text'], 'What is 2 + 2?')
        self.assertEqual(detailed_answers[first_key]['options']['b'], '4')
        self.assertEqual(detailed_answers[first_key]['explanation'], '2 + 2 equals 4.')
        self.assertEqual(detailed_answers[first_key]['topic_tags'], ['arithmetic'])
        self.assertFalse(detailed_answers[second_key]['is_correct'])
        self.assertEqual(attempt.answers, {
            first_key: 'b',
            second_key: 'a',
        })

    def test_feedback_endpoint_returns_attempt_review_summary(self):
        attempt = QuizAttempt.objects.create(
            student=self.student,
            course=self.course,
            exam_paper=self.exam_paper,
            score=1,
            total_questions=2,
            answers={
                str(self.first_question.id): 'b',
                str(self.second_question.id): 'a',
            },
            detailed_answers={
                str(self.first_question.id): {
                    'selected_option': 'b',
                    'correct_option': 'b',
                    'is_correct': True,
                    'question_text': self.first_question.text,
                    'options': self.first_question.available_options,
                    'explanation': self.first_question.explanation,
                    'topic_tags': ['arithmetic'],
                },
                str(self.second_question.id): {
                    'selected_option': 'a',
                    'correct_option': 'c',
                    'is_correct': False,
                    'question_text': self.second_question.text,
                    'options': self.second_question.available_options,
                    'explanation': self.second_question.explanation,
                    'topic_tags': ['functions'],
                },
            },
            mode=QuizAttempt.MODE_PRACTICE,
        )

        response = self.client.get(f'/api/quiz/attempts/{attempt.id}/feedback/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['attempt_id'], attempt.id)
        self.assertEqual(response.data['score'], 1)
        self.assertEqual(response.data['total_questions'], 2)
        self.assertEqual(response.data['percentage'], 50.0)
        self.assertEqual(response.data['summary']['correct'], 1)
        self.assertEqual(response.data['summary']['incorrect'], 1)
        self.assertEqual(response.data['summary']['topics_missed'], [{
            'topic': 'functions',
            'percentage': 0.0,
            'correct': 0,
            'total': 1,
        }])

    def test_feedback_endpoint_is_scoped_to_authenticated_student(self):
        attempt = QuizAttempt.objects.create(
            student=self.other_student,
            score=0,
            total_questions=1,
            answers={str(self.first_question.id): 'a'},
            detailed_answers={},
            mode=QuizAttempt.MODE_PRACTICE,
        )

        response = self.client.get(f'/api/quiz/attempts/{attempt.id}/feedback/')

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data['detail'], 'Attempt not found')

    def test_feedback_endpoint_handles_attempts_without_detailed_answers(self):
        attempt = QuizAttempt.objects.create(
            student=self.student,
            score=0,
            total_questions=1,
            answers={str(self.first_question.id): 'a'},
            mode=QuizAttempt.MODE_PRACTICE,
        )

        response = self.client.get(f'/api/quiz/attempts/{attempt.id}/feedback/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['detailed_answers'], {})
        self.assertEqual(response.data['summary'], {
            'correct': 0,
            'incorrect': 0,
            'topics_missed': [],
        })

    def test_course_topics_returns_unique_topics_for_available_course_questions(self):
        premium_paper = ExamPaper.objects.create(
            title='Premium Programming Quiz',
            course=self.course,
            exam_type=ExamPaper.TYPE_QUIZ,
            year=2027,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )
        Question.objects.create(
            exam_paper=premium_paper,
            text='Premium-only topic question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Premium Topic'],
        )

        response = self.client.get(f'/api/quiz/courses/{self.course.id}/topics/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['course_id'], self.course.id)
        self.assertEqual(response.data['course_name'], self.course.name)
        self.assertEqual(response.data['course_code'], self.course.code)
        self.assertEqual(response.data['chapters'], ['arithmetic', 'functions'])
        self.assertEqual(response.data['total_chapters'], 2)
        self.assertEqual(response.data['topics'], ['arithmetic', 'functions'])
        self.assertEqual(response.data['total_topics'], 2)

    def test_course_chapters_returns_active_chapters_with_question_counts(self):
        chapter = Chapter.objects.create(
            course=self.course,
            number=1,
            title='Introduction',
            description='Foundations',
            icon='📚',
            order=1,
        )
        Chapter.objects.create(
            course=self.course,
            number=2,
            title='Inactive Chapter',
            is_active=False,
        )
        self.first_question.chapter = chapter
        self.first_question.save(update_fields=['chapter'])
        Question.objects.create(
            exam_paper=self.exam_paper,
            chapter=chapter,
            text='Inactive question',
            option_a='A',
            option_b='B',
            correct_option='a',
            is_active=False,
        )

        response = self.client.get(f'/api/quiz/courses/{self.course.id}/chapters/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['course_id'], self.course.id)
        self.assertEqual(response.data['course_name'], self.course.name)
        self.assertEqual(response.data['course_code'], self.course.code)
        self.assertEqual(response.data['total_chapters'], 1)
        self.assertEqual(response.data['chapters'], [
            {
                'id': chapter.id,
                'number': 1,
                'title': 'Introduction',
                'description': 'Foundations',
                'icon': '📚',
                'question_count': 1,
            }
        ])

    def test_course_topics_returns_premium_topics_for_premium_students(self):
        self.student.subscription_status = Student.SUBSCRIPTION_PREMIUM

        self.student.subscription_expiry = timezone.now() + timedelta(days=1)
        self.student.save(update_fields=['subscription_status', 'subscription_expiry'])

        premium_paper = ExamPaper.objects.create(
            title='Premium Programming Quiz',
            course=self.course,
            exam_type=ExamPaper.TYPE_QUIZ,
            year=2027,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )
        Question.objects.create(
            exam_paper=premium_paper,
            text='Premium-only topic question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Premium Topic'],
        )

        response = self.client.get(f'/api/quiz/courses/{self.course.id}/topics/')

        self.assertEqual(
            response.data['topics'],
            ['Premium Topic', 'arithmetic', 'functions'],
        )
        self.assertEqual(
            response.data['chapters'],
            ['arithmetic', 'functions', 'Premium Topic'],
        )

    def test_course_topics_returns_unique_chapters_from_first_topic_tag_only(self):
        Question.objects.create(
            exam_paper=self.exam_paper,
            text='Another Chapter 1 question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Chapter 1: Introduction', 'Definition'],
        )
        Question.objects.create(
            exam_paper=self.exam_paper,
            text='Second Chapter 1 question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Chapter 1: Introduction', 'Basics'],
        )
        Question.objects.create(
            exam_paper=self.exam_paper,
            text='Chapter 2 question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Chapter 2: Risk Analysis', 'Risk'],
        )

        response = self.client.get(f'/api/quiz/courses/{self.course.id}/topics/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['chapters'], [
            'Chapter 1: Introduction',
            'Chapter 2: Risk Analysis',
            'arithmetic',
            'functions',
        ])
        self.assertIn('Definition', response.data['topics'])
        self.assertIn('Basics', response.data['topics'])
        self.assertIn('Risk', response.data['topics'])
        self.assertEqual(
            response.data['chapters_with_count'][:2],
            [
                {
                    'name': 'Chapter 1: Introduction',
                    'question_count': 2,
                },
                {
                    'name': 'Chapter 2: Risk Analysis',
                    'question_count': 1,
                },
            ],
        )

    def test_selective_practice_returns_questions_matching_selected_topics(self):
        response = self.client.post(
            '/api/quiz/selective-practice/',
            {
                'course_id': self.course.id,
                'selected_topics': ['functions'],
                'limit': 10,
            },
            format='json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['course_id'], self.course.id)
        self.assertEqual(response.data['selected_topics'], ['functions'])
        self.assertEqual(response.data['filtered_count'], 1)
        self.assertEqual(response.data['returned_count'], 1)
        self.assertTrue(response.data['can_start'])
        self.assertEqual(response.data['questions'][0]['id'], self.second_question.id)
        self.assertEqual(response.data['questions'][0]['topic_tags'], ['functions'])

    def test_selective_practice_filters_premium_questions_for_free_students(self):
        premium_paper = ExamPaper.objects.create(
            title='Premium Programming Quiz',
            course=self.course,
            exam_type=ExamPaper.TYPE_QUIZ,
            year=2027,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )
        Question.objects.create(
            exam_paper=premium_paper,
            text='Premium-only topic question',
            option_a='A',
            option_b='B',
            correct_option='a',
            topic_tags=['Premium Topic'],
        )

        response = self.client.post(
            '/api/quiz/selective-practice/',
            {
                'course_id': self.course.id,
                'selected_topics': ['Premium Topic'],
            },
            format='json',
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.data['detail'], 'No questions found for selected topics')

    def test_selective_practice_requires_topics(self):
        response = self.client.post(
            '/api/quiz/selective-practice/',
            {
                'course_id': self.course.id,
                'selected_topics': [],
            },
            format='json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data['detail'], 'selected_topics must be a non-empty list')

    def test_selective_attempt_submission_saves_selected_topics(self):
        response = self.client.post(
            '/api/quiz/attempts/',
            {
                'course_id': self.course.id,
                'mode': QuizAttempt.MODE_SELECTIVE,
                'selected_topics': ['functions'],
                'answers': [
                    {
                        'question_id': self.second_question.id,
                        'selected_option': 'c',
                    },
                ],
            },
            format='json',
        )

        self.assertEqual(response.status_code, 201)
        attempt = QuizAttempt.objects.get(id=response.data['attempt_id'])
        self.assertEqual(attempt.mode, QuizAttempt.MODE_SELECTIVE)
        self.assertEqual(attempt.selected_topics, ['functions'])


class QuizFeedbackEngineTests(TestCase):
    def test_get_topics_with_low_score_returns_topics_under_fifty_percent(self):
        weak_topics = get_topics_with_low_score({
            '1': {'is_correct': True, 'topic_tags': ['functions', 'syntax']},
            '2': {'is_correct': False, 'topic_tags': ['functions']},
            '3': {'is_correct': False, 'topic_tags': ['data types']},
        })

        self.assertEqual(weak_topics, [{
            'topic': 'data types',
            'percentage': 0.0,
            'correct': 0,
            'total': 1,
        }])


class QuizChapterImportTests(TestCase):
    def setUp(self):
        self.course = Course.objects.create(
            name='Entrepreneurship',
            code='MGMT221',
        )
        self.exam_paper = ExamPaper.objects.create(
            title='Entrepreneurship Quiz',
            course=self.course,
            exam_type=ExamPaper.TYPE_QUIZ,
            year=2026,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )

    def _build_workbook(self, rows):
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Questions'
        worksheet.append([
            'question_type',
            'question',
            'chapter',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'option_e',
            'correct_option',
            'explanation',
            'difficulty',
            'topic_tags',
            'year_source',
        ])
        worksheet.append([
            'mcq / true_false / fill_blank / matching / essay',
            'Full question text',
            'Required chapter e.g. Chapter 1: Introduction',
            'Option A',
            'Option B',
            'Option C',
            'Option D',
            'Optional option E',
            'a/b/c/d/e',
            'Explanation',
            'easy / medium / hard',
            'Comma separated optional tags',
            'e.g. 2024',
        ])
        for row in rows:
            worksheet.append(row)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def test_excel_import_requires_chapter_and_adds_it_to_topic_tags_first(self):
        workbook = self._build_workbook([
            [
                'mcq',
                'What is entrepreneurship?',
                'Chapter 1: Introduction',
                'Creating value',
                'A license',
                'A tax',
                'A building',
                '',
                'A',
                'Entrepreneurship creates value.',
                'easy',
                'Definition, Basics',
                '2024',
            ],
            [
                'mcq',
                'Which row has no chapter?',
                '',
                'A',
                'B',
                'C',
                'D',
                '',
                'A',
                '',
                'easy',
                'Missing Chapter',
                '2024',
            ],
        ])

        result = import_questions_from_excel(workbook, self.exam_paper)

        self.assertEqual(result.created, 1)
        self.assertEqual(result.skipped, 1)
        self.assertIn('Row 4: skipped — chapter is required.', result.errors)

        question = Question.objects.get()
        self.assertEqual(question.text, 'What is entrepreneurship?')
        self.assertEqual(question.exam_paper, self.exam_paper)
        self.assertEqual(
            question.topic_tags,
            ['Chapter 1: Introduction', 'Definition', 'Basics'],
        )
        self.assertEqual(question.chapter.course, self.course)
        self.assertEqual(question.chapter.number, 1)
        self.assertEqual(question.chapter.title, 'Introduction')
        self.assertFalse(question.is_active)

    def test_excel_import_does_not_duplicate_chapter_when_already_in_tags(self):
        workbook = self._build_workbook([[
            'mcq',
            'What is risk?',
            'Chapter 2: Risk Analysis',
            'Market uncertainty',
            'Free capital',
            'No competition',
            'Guaranteed success',
            '',
            'A',
            '',
            'medium',
            'Chapter 2: Risk Analysis, Chapter 2: Risk Analysis, Risk, Risk',
            '2024',
        ]])

        result = import_questions_from_excel(workbook, self.exam_paper)

        self.assertEqual(result.created, 1)
        question = Question.objects.get()
        self.assertEqual(
            question.topic_tags,
            ['Chapter 2: Risk Analysis', 'Risk'],
        )
        self.assertEqual(question.chapter.number, 2)
        self.assertEqual(question.chapter.title, 'Risk Analysis')

    def test_import_command_validates_exam_paper_course(self):
        other_course = Course.objects.create(
            name='Accounting',
            code='ACFN101',
        )
        workbook = self._build_workbook([])

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(workbook.read())
            tmp.flush()

            with self.assertRaises(CommandError):
                call_command(
                    'import_questions_with_chapters',
                    tmp.name,
                    exam_paper_id=self.exam_paper.id,
                    course_id=other_course.id,
                    stdout=StringIO(),
                )

    def test_import_command_imports_chapter_mapped_questions(self):
        workbook = self._build_workbook([[
            'mcq',
            'What is innovation?',
            'Chapter 1: Introduction',
            'Creating new ideas',
            'Copying others',
            'Waiting',
            'Avoiding change',
            '',
            'A',
            '',
            'easy',
            'Innovation',
            '2024',
        ]])

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(workbook.read())
            tmp.flush()
            output = StringIO()

            call_command(
                'import_questions_with_chapters',
                tmp.name,
                exam_paper_id=self.exam_paper.id,
                course_id=self.course.id,
                stdout=output,
            )

        self.assertEqual(Question.objects.count(), 1)
        question = Question.objects.get()
        self.assertEqual(question.chapter.number, 1)
        self.assertEqual(question.chapter.title, 'Introduction')
        self.assertIn('Created: 1', output.getvalue())

    def test_import_course_with_chapters_command_creates_chapters_and_questions(self):
        from openpyxl import Workbook

        workbook = Workbook()
        chapters = workbook.active
        chapters.title = 'CHAPTERS'
        chapters.append(['number', 'title', 'description', 'icon'])
        chapters.append([1, 'Introduction', 'Foundations', '📚'])
        questions = workbook.create_sheet('QUESTIONS')
        questions.append([
            'question_text',
            'chapter_number',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'explanation',
            'difficulty',
            'topic_tags',
            'year_source',
        ])
        questions.append([
            'What is entrepreneurship?',
            1,
            'Creating value',
            'A license',
            'A tax',
            'A building',
            'a',
            'Entrepreneurship creates value.',
            'easy',
            'Definition, Basics',
            '2024',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(output.read())
            tmp.flush()
            command_output = StringIO()

            call_command(
                'import_course_with_chapters',
                file=tmp.name,
                course_id=self.course.id,
                exam_paper_id=self.exam_paper.id,
                stdout=command_output,
            )

        chapter = Chapter.objects.get(course=self.course, number=1)
        question = Question.objects.get()
        self.assertEqual(chapter.title, 'Introduction')
        self.assertEqual(question.chapter, chapter)
        self.assertEqual(question.exam_paper, self.exam_paper)
        self.assertTrue(question.is_active)
        self.assertEqual(
            question.topic_tags,
            ['Chapter 1: Introduction', 'Definition', 'Basics'],
        )
        self.assertIn('Questions created: 1', command_output.getvalue())

    def test_import_course_with_chapters_command_auto_extracts_chapters_from_single_sheet(self):
        from openpyxl import Workbook

        workbook = Workbook()
        questions = workbook.active
        questions.title = 'Generated'
        questions.append([
            'chapter',
            'question',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'explanation',
            'difficulty',
            'topic_tags',
            'year_source',
        ])
        questions.append([
            'Chapter 1: Introduction to Arrays',
            'What is an array?',
            'An ordered collection',
            'A function type',
            'A storage device',
            'Network protocol',
            'a',
            'Arrays store elements consecutively.',
            'easy',
            'array,basics',
            '2023',
        ])
        questions.append([
            '1: Introduction to Arrays',
            'How are arrays stored in memory?',
            'Consecutively',
            'In a linked list',
            'In a tree',
            'In a hash table',
            'a',
            'Arrays allocate contiguous memory.',
            'medium',
            'array,memory',
            '2023',
        ])
        questions.append([
            'Chapter 2',
            'What is a linked list?',
            'List with pointers',
            'Array with links',
            'Tree structure',
            'Graph structure',
            'a',
            'Linked lists use node pointers.',
            'medium',
            'linked-list,basics',
            '',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(output.read())
            tmp.flush()
            command_output = StringIO()

            call_command(
                'import_course_with_chapters',
                file=tmp.name,
                course_id=self.course.id,
                stdout=command_output,
            )

        self.assertEqual(Chapter.objects.count(), 2)
        self.assertEqual(Question.objects.count(), 3)

        first_chapter = Chapter.objects.get(course=self.course, number=1)
        second_chapter = Chapter.objects.get(course=self.course, number=2)
        self.assertEqual(first_chapter.title, 'Introduction to Arrays')
        self.assertEqual(second_chapter.title, 'Chapter 2')

        first_question = Question.objects.get(text='What is an array?')
        self.assertEqual(first_question.chapter, first_chapter)
        self.assertEqual(
            first_question.topic_tags,
            ['Chapter 1: Introduction to Arrays', 'array', 'basics'],
        )
        self.assertIn('Format: single sheet', command_output.getvalue())
        self.assertIn('Questions created: 3', command_output.getvalue())

    def test_import_exit_exam_questions_sets_exam_course_and_leaves_chapter_empty(self):
        from openpyxl import Workbook

        department = Department.objects.create(name='Computing')
        exit_paper = ExamPaper.objects.create(
            title='Computing Exit Model',
            department=department,
            exam_type=ExamPaper.TYPE_EXIT_MODEL,
            year=2026,
            duration_minutes=120,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )
        workbook = Workbook()
        questions = workbook.active
        questions.title = 'Questions'
        questions.append([
            'course',
            'question',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'explanation',
            'difficulty',
            'topic_tags',
            'year_source',
        ])
        questions.append([
            'Required course name or code',
            'Full question text',
            'Option A',
            'Option B',
            'Option C',
            'Option D',
            'a/b/c/d/e',
            'Explanation',
            'easy / medium / hard',
            'Comma separated optional tags',
            'e.g. 2024',
        ])
        questions.append([
            self.course.name,
            'What is entrepreneurship?',
            'Creating value',
            'A license',
            'A tax',
            'A building',
            'a',
            'Entrepreneurship creates value.',
            'easy',
            'Definition, Basics',
            '2024',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(output.read())
            tmp.flush()
            command_output = StringIO()

            call_command(
                'import_exit_exam_questions',
                file=tmp.name,
                exam_paper_id=exit_paper.id,
                stdout=command_output,
            )

        exit_paper.refresh_from_db()
        question = Question.objects.get()
        self.assertEqual(exit_paper.course, self.course)
        self.assertEqual(question.exam_paper, exit_paper)
        self.assertIsNone(question.chapter)
        self.assertTrue(question.is_active)
        self.assertEqual(question.topic_tags, ['Definition', 'Basics'])
        self.assertEqual(question.year_source, '2024')
        self.assertIn('Questions created: 1', command_output.getvalue())

    def test_import_exit_exam_questions_rejects_multiple_courses_for_one_paper(self):
        from openpyxl import Workbook

        department = Department.objects.create(name='Computing')
        other_course = Course.objects.create(
            name='Data Structures',
            code='CS202',
        )
        exit_paper = ExamPaper.objects.create(
            title='Computing Exit Model',
            department=department,
            exam_type=ExamPaper.TYPE_EXIT_MODEL,
            year=2026,
            duration_minutes=120,
            access_level=ExamPaper.ACCESS_PREMIUM,
        )
        workbook = Workbook()
        questions = workbook.active
        questions.append([
            'course',
            'question',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'explanation',
        ])
        for course_name in [self.course.name, other_course.name]:
            questions.append([
                course_name,
                f'Question for {course_name}?',
                'A',
                'B',
                'C',
                'D',
                'a',
                'Because A is correct.',
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
            tmp.write(output.read())
            tmp.flush()

            with self.assertRaises(CommandError):
                call_command(
                    'import_exit_exam_questions',
                    file=tmp.name,
                    exam_paper_id=exit_paper.id,
                    stdout=StringIO(),
                )

        self.assertEqual(Question.objects.count(), 0)

    def test_admin_import_page_uses_exit_template_for_exit_papers(self):
        department = Department.objects.create(name='Computing')
        exit_paper = ExamPaper.objects.create(
            title='Computing Exit Official',
            department=department,
            exam_type=ExamPaper.TYPE_EXIT_REAL,
            year=2026,
            duration_minutes=120,
        )
        model_admin = ExamPaperAdmin(ExamPaper, django_admin.site)

        exit_template = model_admin._import_template_context(exit_paper)
        regular_template = model_admin._import_template_context(self.exam_paper)

        self.assertEqual(exit_template['filename'], 'exit_exam_import_template.xlsx')
        self.assertEqual(regular_template['filename'], 'quiz_import_template.xlsx')

    def test_admin_exit_excel_importer_accepts_course_template(self):
        from openpyxl import Workbook

        department = Department.objects.create(name='Computing')
        exit_paper = ExamPaper.objects.create(
            title='Computing Exit Official',
            department=department,
            exam_type=ExamPaper.TYPE_EXIT_REAL,
            year=2026,
            duration_minutes=120,
        )
        workbook = Workbook()
        questions = workbook.active
        questions.title = 'Questions'
        questions.append([
            'course',
            'question',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'option_e',
            'correct_option',
            'explanation',
            'difficulty',
            'topic_tags',
            'year_source',
        ])
        questions.append([
            'Required course name or code',
            'Full question text',
            'Option A',
            'Option B',
            'Option C',
            'Option D',
            'Optional option E',
            'a/b/c/d/e',
            'Explanation',
            'easy / medium / hard',
            'Comma separated optional tags',
            'e.g. 2024',
        ])
        questions.append([
            self.course.code,
            'What is an exit exam?',
            'A comprehensive assessment',
            'A chapter quiz',
            'A license',
            'A form',
            '',
            'a',
            'Exit exams assess broader course knowledge.',
            'medium',
            'assessment,basics',
            '2026',
        ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        result = import_exit_questions_from_excel(output, exit_paper)

        exit_paper.refresh_from_db()
        question = Question.objects.get()
        self.assertEqual(result.created, 1)
        self.assertEqual(result.skipped, 0)
        self.assertEqual(exit_paper.course, self.course)
        self.assertEqual(question.exam_paper, exit_paper)
        self.assertIsNone(question.chapter)
        self.assertFalse(question.is_active)
        self.assertEqual(question.topic_tags, ['assessment', 'basics'])
