"""
apps/quiz/importers.py

Handles importing questions from an Excel file into a given ExamPaper.
Used by the admin upload action on ExamPaper.
"""

import logging
import re
from typing import NamedTuple

from django.db.models import Max, Q

logger = logging.getLogger(__name__)

VALID_TYPES = {'mcq', 'true_false', 'fill_blank', 'matching', 'essay'}
VALID_DIFFICULTIES = {'easy', 'medium', 'hard'}
VALID_OPTIONS = {'a', 'b', 'c', 'd', 'e'}


class ImportResult(NamedTuple):
    created: int
    skipped: int
    errors: list[str]


def parse_chapter_label(value: str) -> tuple[int | None, str]:
    text = str(value or "").strip()
    match = re.match(r'^(?:chapter\s*)?(\d+)\s*[:.)\-]?\s*(.*)$', text, re.IGNORECASE)
    if not match:
        return None, text

    number = int(match.group(1))
    title = match.group(2).strip() or f'Chapter {number}'
    return number, title


def get_or_create_chapter_from_label(course, label):
    if not course:
        return None

    number, title = parse_chapter_label(label)
    from apps.quiz.models import Chapter

    if number is not None:
        chapter, created = Chapter.objects.get_or_create(
            course=course,
            number=number,
            defaults={
                'title': title,
                'order': number,
                'is_active': True,
            },
        )
        if not created and title and chapter.title != title:
            chapter.title = title
            chapter.save(update_fields=['title', 'updated_at'])
        return chapter

    chapter = Chapter.objects.filter(course=course, title__iexact=title).first()
    if chapter:
        return chapter

    next_number = (
        Chapter.objects.filter(course=course).aggregate(max_number=Max('number'))['max_number']
        or 0
    ) + 1
    return Chapter.objects.create(
        course=course,
        number=next_number,
        title=title,
        order=next_number,
        is_active=True,
    )


def import_questions_from_excel(file, exam_paper) -> ImportResult:
    """
    Reads an uploaded Excel file and creates Question records
    linked to the given ExamPaper.

    Returns ImportResult(created, skipped, errors).
    """
    try:
        import openpyxl
    except ImportError:
        return ImportResult(0, 0, ["openpyxl is not installed. Run: pip install openpyxl"])

    from apps.quiz.models import Question

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return ImportResult(0, 0, [f"Could not open Excel file: {e}"])

    # Support both 'Questions' sheet name and first sheet
    if "Questions" in wb.sheetnames:
        ws = wb["Questions"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        return ImportResult(0, 0, ["File has no data rows. Make sure you start from row 3."])

    # Row 1 = headers, Row 2 = notes, Row 3+ = data
    header_row = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try:
            return header_row.index(name)
        except ValueError:
            return None

    # Map column indices
    COL = {
        "question_type":  col("question_type"),
        "question":       col("question"),
        "chapter":        col("chapter"),
        "option_a":       col("option_a"),
        "option_b":       col("option_b"),
        "option_c":       col("option_c"),
        "option_d":       col("option_d"),
        "option_e":       col("option_e"),
        "correct_option": col("correct_option"),
        "explanation":    col("explanation"),
        "difficulty":     col("difficulty"),
        "topic_tags":     col("topic_tags"),
        "year_source":    col("year_source"),
    }

    if COL["question_type"] is None or COL["question"] is None or COL["chapter"] is None:
        return ImportResult(0, 0, [
            "Could not find required columns 'question_type', 'question', and 'chapter'. "
            "Make sure you are using the official template."
        ])

    def get(row, key):
        idx = COL.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    created = 0
    skipped = 0
    errors = []
    year_source_max = Question._meta.get_field("year_source").max_length

    def normalize_year_source(value):
        value = (value or "").strip()
        if not value or len(value) > year_source_max:
            return ""
        return value

    data_rows = rows[2:]  # skip header + notes

    for row_num, row in enumerate(data_rows, start=3):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        q_type = get(row, "question_type").lower()
        q_text = get(row, "question")
        chapter = get(row, "chapter")

        if not q_text:
            errors.append(f"Row {row_num}: skipped — question text is empty.")
            skipped += 1
            continue

        if not chapter:
            errors.append(f"Row {row_num}: skipped — chapter is required.")
            skipped += 1
            continue

        if q_type not in VALID_TYPES:
            errors.append(
                f"Row {row_num}: skipped — unknown question_type '{q_type}'. "
                f"Must be one of: {', '.join(sorted(VALID_TYPES))}."
            )
            skipped += 1
            continue

        # Options
        option_a = get(row, "option_a")
        option_b = get(row, "option_b")
        option_c = get(row, "option_c")
        option_d = get(row, "option_d")
        option_e = get(row, "option_e")
        correct  = get(row, "correct_option")

        # Enforce true_false options
        if q_type == "true_false":
            option_a = "True"
            option_b = "False"
            option_c = option_d = option_e = ""
            if correct.lower() not in ("a", "b", "true", "false", ""):
                correct = ""
            elif correct.lower() == "true":
                correct = "a"
            elif correct.lower() == "false":
                correct = "b"

        # Normalize correct_option for MCQ/true_false
        if q_type in ("mcq", "true_false"):
            correct = correct.lower()
            if correct not in VALID_OPTIONS:
                correct = ""

        # Essay/matching — no correct option
        if q_type in ("essay", "matching"):
            correct = ""

        # Difficulty
        difficulty = get(row, "difficulty").lower()
        if difficulty not in VALID_DIFFICULTIES:
            difficulty = "medium"

        # Topic tags
        raw_tags = get(row, "topic_tags")
        if raw_tags:
            raw_topic_tags = [t.strip() for t in raw_tags.split(",") if t.strip()]
        else:
            raw_topic_tags = []

        topic_tags = []
        seen_tags = set()
        for tag in [chapter, *raw_topic_tags]:
            if tag in seen_tags:
                continue
            seen_tags.add(tag)
            topic_tags.append(tag)

        # Explanation
        explanation = get(row, "explanation")

        try:
            year_source = normalize_year_source(get(row, "year_source"))
            chapter_obj = get_or_create_chapter_from_label(exam_paper.course, chapter)

            Question.objects.create(
                exam_paper=exam_paper,
                chapter=chapter_obj,
                question_type=q_type,
                text=q_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                option_e=option_e,
                correct_option=correct,
                explanation=explanation,
                difficulty=difficulty,
                topic_tags=topic_tags,
                year_source=year_source,
                is_active=False,  # admin reviews before activating
            )
            created += 1
        except Exception as e:
            errors.append(f"Row {row_num}: failed to save — {e}")
            skipped += 1

    return ImportResult(created=created, skipped=skipped, errors=errors)


def import_exit_questions_from_excel(file, exam_paper) -> ImportResult:
    """
    Reads an uploaded exit-exam Excel file and creates Question records linked
    to the given ExamPaper using a course column instead of a chapter column.
    """
    try:
        import openpyxl
    except ImportError:
        return ImportResult(0, 0, ["openpyxl is not installed. Run: pip install openpyxl"])

    from apps.content.models import Course
    from apps.quiz.models import Question

    if not exam_paper.is_exit_exam:
        return ImportResult(0, 0, ["This importer only supports exit exam papers."])

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return ImportResult(0, 0, [f"Could not open Excel file: {e}"])

    ws = wb["Questions"] if "Questions" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        return ImportResult(0, 0, ["File has no data rows. Make sure you start from row 3."])

    header_row = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(name):
        try:
            return header_row.index(name)
        except ValueError:
            return None

    COL = {
        "question_type": col("question_type"),
        "course": col("course"),
        "question": col("question"),
        "question_text": col("question_text"),
        "option_a": col("option_a"),
        "option_b": col("option_b"),
        "option_c": col("option_c"),
        "option_d": col("option_d"),
        "option_e": col("option_e"),
        "correct_option": col("correct_option"),
        "explanation": col("explanation"),
        "difficulty": col("difficulty"),
        "topic_tags": col("topic_tags"),
        "year_source": col("year_source"),
    }

    if COL["course"] is None or (COL["question"] is None and COL["question_text"] is None):
        return ImportResult(0, 0, [
            "Could not find required columns 'course' and 'question'. "
            "Make sure you are using the official exit exam template."
        ])

    def get(row, key):
        idx = COL.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    def resolve_course(value):
        value = (value or "").strip()
        return Course.objects.filter(
            Q(name__iexact=value) | Q(code__iexact=value),
            is_active=True,
        ).first()

    created = 0
    skipped = 0
    errors = []
    resolved_courses = {}
    year_source_max = Question._meta.get_field("year_source").max_length

    for row_num, row in enumerate(rows[2:], start=3):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        q_type = get(row, "question_type").lower() or Question.TYPE_MCQ
        q_text = get(row, "question") or get(row, "question_text")
        course_label = get(row, "course")

        if not course_label:
            errors.append(f"Row {row_num}: skipped — course is required.")
            skipped += 1
            continue

        course = resolve_course(course_label)
        if course is None:
            errors.append(f"Row {row_num}: skipped — active course '{course_label}' was not found.")
            skipped += 1
            continue

        if exam_paper.course_id and exam_paper.course_id != course.id:
            errors.append(
                f"Row {row_num}: skipped — course '{course_label}' does not match "
                f"exam paper course '{exam_paper.course.name}'."
            )
            skipped += 1
            continue

        if not q_text:
            errors.append(f"Row {row_num}: skipped — question text is empty.")
            skipped += 1
            continue

        if q_type not in VALID_TYPES:
            errors.append(
                f"Row {row_num}: skipped — unknown question_type '{q_type}'. "
                f"Must be one of: {', '.join(sorted(VALID_TYPES))}."
            )
            skipped += 1
            continue

        option_a = get(row, "option_a")
        option_b = get(row, "option_b")
        option_c = get(row, "option_c")
        option_d = get(row, "option_d")
        option_e = get(row, "option_e")
        correct = get(row, "correct_option")

        if q_type == "true_false":
            option_a = "True"
            option_b = "False"
            option_c = option_d = option_e = ""
            if correct.lower() == "true":
                correct = "a"
            elif correct.lower() == "false":
                correct = "b"

        if q_type in ("mcq", "true_false"):
            correct = correct.lower()
            if correct not in VALID_OPTIONS:
                errors.append(f"Row {row_num}: skipped — invalid correct_option '{correct}'.")
                skipped += 1
                continue

        if q_type in ("essay", "matching"):
            correct = ""

        difficulty = get(row, "difficulty").lower()
        if difficulty not in VALID_DIFFICULTIES:
            difficulty = "medium"

        raw_tags = get(row, "topic_tags")
        topic_tags = []
        seen_tags = set()
        if raw_tags:
            for tag in raw_tags.split(","):
                tag = tag.strip()
                if tag and tag not in seen_tags:
                    seen_tags.add(tag)
                    topic_tags.append(tag)

        year_source = get(row, "year_source")
        if not year_source or len(year_source) > year_source_max:
            year_source = ""

        try:
            if exam_paper.course_id is None:
                resolved_courses[course.id] = course
                if len(resolved_courses) > 1:
                    errors.append(
                        f"Row {row_num}: skipped — one exit exam paper can only import one course."
                    )
                    skipped += 1
                    continue

            Question.objects.create(
                exam_paper=exam_paper,
                chapter=None,
                question_type=q_type,
                text=q_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                option_e=option_e,
                correct_option=correct,
                explanation=get(row, "explanation"),
                difficulty=difficulty,
                topic_tags=topic_tags,
                year_source=year_source,
                is_active=False,
            )
            created += 1
        except Exception as e:
            errors.append(f"Row {row_num}: failed to save — {e}")
            skipped += 1

    if created and exam_paper.course_id is None and len(resolved_courses) == 1:
        exam_paper.course = next(iter(resolved_courses.values()))
        exam_paper.save(update_fields=["course", "updated_at"])

    return ImportResult(created=created, skipped=skipped, errors=errors)
