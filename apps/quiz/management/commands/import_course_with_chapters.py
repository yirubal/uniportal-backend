from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

from apps.content.models import Course
from apps.quiz.importers import parse_chapter_label
from apps.quiz.models import Chapter, ExamPaper, Question


VALID_TYPES = {'mcq', 'true_false', 'fill_blank', 'matching', 'essay'}
VALID_DIFFICULTIES = {'easy', 'medium', 'hard'}
VALID_OPTIONS = {'a', 'b', 'c', 'd', 'e'}


class Command(BaseCommand):
    help = 'Import chapters and questions from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to Excel file with one QUESTIONS sheet or CHAPTERS + QUESTIONS sheets',
        )
        parser.add_argument(
            '--course-id',
            type=int,
            required=True,
            help='Course ID to import into',
        )
        parser.add_argument(
            '--exam-paper-id',
            type=int,
            default=None,
            help='Optional ExamPaper ID to link imported questions to',
        )
        parser.add_argument(
            '--delete-existing',
            action='store_true',
            help='Delete existing chapters and questions for this course before importing',
        )

    def handle(self, *args, **options):
        course = self._get_course(options['course_id'])
        exam_paper = self._get_exam_paper(options.get('exam_paper_id'), course)
        workbook = self._load_workbook(options['file'])

        sheets = {name.upper(): workbook[name] for name in workbook.sheetnames}
        has_two_sheet_format = 'CHAPTERS' in sheets and 'QUESTIONS' in sheets
        questions_sheet = sheets.get('QUESTIONS') or workbook[workbook.sheetnames[0]]

        with transaction.atomic():
            if options.get('delete_existing'):
                self._delete_existing(course)

            if has_two_sheet_format:
                self.stdout.write('Format: two sheets (CHAPTERS + QUESTIONS)')
                chapters = self._create_chapters(course, sheets['CHAPTERS'])
            else:
                self.stdout.write('Format: single sheet (auto-extracting chapters)')
                chapters = self._auto_extract_chapters(course, questions_sheet)

            created_count, skipped_count, errors = self._create_questions(
                course=course,
                questions_sheet=questions_sheet,
                chapters=chapters,
                exam_paper=exam_paper,
            )

        self.stdout.write(self.style.SUCCESS('Import complete'))
        self.stdout.write(f'  Course: {course.code} - {course.name}')
        self.stdout.write(f'  Chapters: {len(chapters)}')
        self.stdout.write(f'  Questions created: {created_count}')
        self.stdout.write(f'  Questions skipped: {skipped_count}')

        if errors:
            self.stdout.write(self.style.WARNING('Errors:'))
            for error in errors[:20]:
                self.stdout.write(f'  {error}')
            if len(errors) > 20:
                self.stdout.write(f'  ... and {len(errors) - 20} more')

        if created_count == 0 and errors:
            raise CommandError('No questions were imported.')

    def _get_course(self, course_id):
        try:
            return Course.objects.get(id=course_id, is_active=True)
        except Course.DoesNotExist as exc:
            raise CommandError(f'Active course not found: {course_id}') from exc

    def _get_exam_paper(self, exam_paper_id, course):
        if not exam_paper_id:
            return None

        try:
            exam_paper = ExamPaper.objects.get(id=exam_paper_id)
        except ExamPaper.DoesNotExist as exc:
            raise CommandError(f'Exam paper not found: {exam_paper_id}') from exc

        if exam_paper.course_id != course.id:
            raise CommandError(
                f'Exam paper {exam_paper.id} is not linked to course {course.id}.'
            )
        return exam_paper

    def _load_workbook(self, file_path):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl') from exc

        try:
            return openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f'File not found: {file_path}') from exc
        except Exception as exc:
            raise CommandError(f'Could not open Excel file: {exc}') from exc

    def _delete_existing(self, course):
        deleted_questions, _ = Question.objects.filter(
            Q(chapter__course=course) | Q(exam_paper__course=course)
        ).delete()
        deleted_chapters, _ = Chapter.objects.filter(course=course).delete()
        self.stdout.write(
            f'Deleted {deleted_chapters} chapter(s) and {deleted_questions} question(s).'
        )

    def _create_chapters(self, course, chapters_sheet):
        chapters = {}
        for row_num, row in self._iter_dict_rows(chapters_sheet):
            number = self._as_int(row.get('number'))
            title = self._clean(row.get('title'))

            if not number or number < 1:
                raise CommandError(f'CHAPTERS row {row_num}: number must be a positive integer.')
            if not title:
                raise CommandError(f'CHAPTERS row {row_num}: title is required.')

            chapter, _ = Chapter.objects.update_or_create(
                course=course,
                number=number,
                defaults={
                    'title': title,
                    'description': self._clean(row.get('description')) or None,
                    'icon': self._clean(row.get('icon')) or None,
                    'order': self._as_int(row.get('order')) or number,
                    'is_active': self._as_bool(row.get('is_active'), default=True),
                },
            )
            chapters[number] = chapter
            self.stdout.write(f'Chapter {number}: {chapter.title}')

        return chapters

    def _auto_extract_chapters(self, course, questions_sheet):
        headers = self._headers(questions_sheet)
        if 'chapter' not in headers:
            raise CommandError('Questions sheet must have a "chapter" column.')

        chapters = {}
        for row_num, row in self._iter_dict_rows(questions_sheet):
            chapter_label = self._clean(row.get('chapter'))
            if not chapter_label:
                continue

            number, title = parse_chapter_label(chapter_label)
            if number is None:
                self.stdout.write(
                    self.style.WARNING(
                        f'QUESTIONS row {row_num}: could not extract chapter number from "{chapter_label}".'
                    )
                )
                continue

            if number in chapters:
                continue

            chapter, created = Chapter.objects.update_or_create(
                course=course,
                number=number,
                defaults={
                    'title': title,
                    'order': number,
                    'is_active': True,
                },
            )
            chapters[number] = chapter
            status = 'Created' if created else 'Updated'
            self.stdout.write(f'{status} chapter {number}: {chapter.title}')

        if not chapters:
            raise CommandError('No valid chapters could be extracted from the questions sheet.')

        return chapters

    def _create_questions(self, course, questions_sheet, chapters, exam_paper):
        created = 0
        skipped = 0
        errors = []

        for row_num, row in self._iter_dict_rows(questions_sheet):
            chapter = self._resolve_question_chapter(course, row, chapters)
            if chapter is None:
                errors.append(f'QUESTIONS row {row_num}: chapter_number or chapter is required.')
                skipped += 1
                continue

            question_text = self._clean(row.get('question_text')) or self._clean(row.get('question'))
            if not question_text:
                errors.append(f'QUESTIONS row {row_num}: question_text is required.')
                skipped += 1
                continue

            missing_options = [
                name
                for name in ['option_a', 'option_b', 'option_c', 'option_d']
                if not self._clean(row.get(name))
            ]
            if missing_options:
                errors.append(
                    f'QUESTIONS row {row_num}: missing required option(s): {", ".join(missing_options)}.'
                )
                skipped += 1
                continue

            correct_option = self._clean(row.get('correct_option')).lower()
            if correct_option and correct_option not in VALID_OPTIONS:
                errors.append(f'QUESTIONS row {row_num}: invalid correct_option "{correct_option}".')
                skipped += 1
                continue
            if not correct_option:
                errors.append(f'QUESTIONS row {row_num}: correct_option is required.')
                skipped += 1
                continue

            explanation = self._clean(row.get('explanation'))
            if not explanation:
                errors.append(f'QUESTIONS row {row_num}: explanation is required.')
                skipped += 1
                continue

            difficulty = self._clean(row.get('difficulty')).lower() or 'medium'
            if difficulty not in VALID_DIFFICULTIES:
                difficulty = 'medium'

            question_type = self._clean(row.get('question_type')).lower() or Question.TYPE_MCQ
            if question_type not in VALID_TYPES:
                errors.append(f'QUESTIONS row {row_num}: invalid question_type "{question_type}".')
                skipped += 1
                continue

            topic_tags = self._topic_tags(chapter, row.get('topic_tags'))

            Question.objects.create(
                exam_paper=exam_paper,
                chapter=chapter,
                question_type=question_type,
                text=question_text,
                option_a=self._clean(row.get('option_a')),
                option_b=self._clean(row.get('option_b')),
                option_c=self._clean(row.get('option_c')),
                option_d=self._clean(row.get('option_d')),
                option_e=self._clean(row.get('option_e')),
                correct_option=correct_option,
                explanation=explanation,
                difficulty=difficulty,
                topic_tags=topic_tags,
                year_source=self._year_source(row.get('year_source')),
                is_active=True,
            )
            created += 1

        return created, skipped, errors

    def _resolve_question_chapter(self, course, row, chapters):
        chapter_number = self._as_int(row.get('chapter_number')) or self._as_int(row.get('chapter#'))
        if chapter_number:
            return chapters.get(chapter_number)

        chapter_label = self._clean(row.get('chapter'))
        if not chapter_label:
            return None

        chapter_number, _title = parse_chapter_label(chapter_label)
        if chapter_number is None:
            return None

        return chapters.get(chapter_number)

    def _iter_dict_rows(self, worksheet):
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return

        keys = [self._clean(value).lower() for value in header]
        for row_num, values in enumerate(rows, start=2):
            if all(self._clean(value) == '' for value in values):
                continue
            yield row_num, {
                key: values[index] if index < len(values) else None
                for index, key in enumerate(keys)
                if key
            }

    def _headers(self, worksheet):
        first_row = next(worksheet.iter_rows(max_row=1, values_only=True), None)
        if not first_row:
            return set()
        return {self._clean(value).lower() for value in first_row if self._clean(value)}

    def _topic_tags(self, chapter, raw_tags):
        tags = [f'Chapter {chapter.number}: {chapter.title}']
        tags.extend(
            tag.strip()
            for tag in self._clean(raw_tags).split(',')
            if tag.strip()
        )

        deduped = []
        seen = set()
        for tag in tags:
            if tag in seen:
                continue
            seen.add(tag)
            deduped.append(tag)
        return deduped

    @staticmethod
    def _clean(value):
        return str(value).strip() if value is not None else ''

    @staticmethod
    def _as_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _as_bool(value, default):
        if value is None or str(value).strip() == '':
            return default
        return str(value).strip().lower() in {'1', 'true', 'yes', 'y', 'active'}

    @staticmethod
    def _year_source(value):
        value = str(value).strip() if value is not None else ''
        field = Question._meta.get_field('year_source')
        if not value or len(value) > field.max_length:
            return ''
        return value
