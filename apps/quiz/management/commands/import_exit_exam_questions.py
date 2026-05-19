from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q

from apps.content.models import Course
from apps.quiz.models import ExamPaper, Question


VALID_TYPES = {'mcq', 'true_false', 'fill_blank', 'matching', 'essay'}
VALID_DIFFICULTIES = {'easy', 'medium', 'hard'}
VALID_OPTIONS = {'a', 'b', 'c', 'd', 'e'}


class Command(BaseCommand):
    help = 'Import exit exam questions from an Excel file using course-based filtering'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to Excel file with exit exam questions',
        )
        parser.add_argument(
            '--exam-paper-id',
            type=int,
            required=True,
            help='Exit ExamPaper ID to link imported questions to',
        )
        parser.add_argument(
            '--delete-existing',
            action='store_true',
            help='Delete existing questions for this exam paper before importing',
        )

    def handle(self, *args, **options):
        exam_paper = self._get_exit_exam_paper(options['exam_paper_id'])
        workbook = self._load_workbook(options['file'])
        questions_sheet = self._get_questions_sheet(workbook)

        self.stdout.write(f'Importing exit exam questions for: {exam_paper.title}')

        with transaction.atomic():
            rows, skipped, errors, course = self._prepare_questions(exam_paper, questions_sheet)

            if not rows:
                self._write_errors(errors)
                raise CommandError('No questions were imported.')

            if exam_paper.course_id is None:
                exam_paper.course = course
                exam_paper.save(update_fields=['course', 'updated_at'])
                self.stdout.write(f'Linked exam paper to course: {course.code} - {course.name}')

            if options.get('delete_existing'):
                deleted, _ = exam_paper.questions.all().delete()
                self.stdout.write(f'Deleted {deleted} existing question(s).')

            created = self._create_questions(exam_paper, rows)

        self.stdout.write(self.style.SUCCESS('Import complete'))
        self.stdout.write(f'  Exam: {exam_paper.title}')
        self.stdout.write(f'  Course: {exam_paper.course.code} - {exam_paper.course.name}')
        self.stdout.write(f'  Questions created: {created}')
        self.stdout.write(f'  Questions skipped: {skipped}')
        self._write_errors(errors)

    def _get_exit_exam_paper(self, exam_paper_id):
        try:
            exam_paper = ExamPaper.objects.select_related('course').get(id=exam_paper_id)
        except ExamPaper.DoesNotExist as exc:
            raise CommandError(f'Exam paper not found: {exam_paper_id}') from exc

        if not exam_paper.is_exit_exam:
            raise CommandError(
                f'Exam paper {exam_paper.id} is not an exit exam. '
                f'Use exam_type "{ExamPaper.TYPE_EXIT_REAL}" or "{ExamPaper.TYPE_EXIT_MODEL}".'
            )
        return exam_paper

    def _load_workbook(self, file_path):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl') from exc

        try:
            return openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f'File not found: {file_path}') from exc
        except Exception as exc:
            raise CommandError(f'Could not open Excel file: {exc}') from exc

    @staticmethod
    def _get_questions_sheet(workbook):
        sheets = {name.upper(): workbook[name] for name in workbook.sheetnames}
        return sheets.get('QUESTIONS') or workbook[workbook.sheetnames[0]]

    def _prepare_questions(self, exam_paper, worksheet):
        headers = self._headers(worksheet)
        if 'course' not in headers:
            raise CommandError('Questions sheet must have a "course" column.')
        if 'question' not in headers and 'question_text' not in headers:
            raise CommandError('Questions sheet must have a "question" column.')

        prepared = []
        skipped = 0
        errors = []
        courses_by_id = {}

        for row_num, row in self._iter_dict_rows(worksheet):
            if self._is_notes_row(row):
                continue

            course_label = self._clean(row.get('course'))
            if not course_label:
                errors.append(f'QUESTIONS row {row_num}: course is required.')
                skipped += 1
                continue

            course = self._resolve_course(course_label)
            if course is None:
                errors.append(f'QUESTIONS row {row_num}: active course "{course_label}" not found.')
                skipped += 1
                continue

            if exam_paper.course_id and exam_paper.course_id != course.id:
                errors.append(
                    f'QUESTIONS row {row_num}: course "{course_label}" does not match '
                    f'exam paper course "{exam_paper.course.name}".'
                )
                skipped += 1
                continue

            question_text = self._clean(row.get('question')) or self._clean(row.get('question_text'))
            if not question_text:
                errors.append(f'QUESTIONS row {row_num}: question is required.')
                skipped += 1
                continue

            missing_options = [
                name
                for name in ['option_a', 'option_b', 'option_c', 'option_d']
                if not self._clean(row.get(name))
            ]
            if missing_options:
                errors.append(
                    f'QUESTIONS row {row_num}: missing required option(s): {", ".join(missing_options)}.'
                )
                skipped += 1
                continue

            correct_option = self._clean(row.get('correct_option')).lower()
            if correct_option not in VALID_OPTIONS:
                errors.append(f'QUESTIONS row {row_num}: invalid correct_option "{correct_option}".')
                skipped += 1
                continue

            explanation = self._clean(row.get('explanation'))
            if not explanation:
                errors.append(f'QUESTIONS row {row_num}: explanation is required.')
                skipped += 1
                continue

            question_type = self._clean(row.get('question_type')).lower() or Question.TYPE_MCQ
            if question_type not in VALID_TYPES:
                errors.append(f'QUESTIONS row {row_num}: invalid question_type "{question_type}".')
                skipped += 1
                continue

            difficulty = self._clean(row.get('difficulty')).lower() or Question.DIFFICULTY_MEDIUM
            if difficulty not in VALID_DIFFICULTIES:
                difficulty = Question.DIFFICULTY_MEDIUM

            courses_by_id[course.id] = course
            prepared.append({
                'question_type': question_type,
                'text': question_text,
                'option_a': self._clean(row.get('option_a')),
                'option_b': self._clean(row.get('option_b')),
                'option_c': self._clean(row.get('option_c')),
                'option_d': self._clean(row.get('option_d')),
                'option_e': self._clean(row.get('option_e')),
                'correct_option': correct_option,
                'explanation': explanation,
                'difficulty': difficulty,
                'topic_tags': self._topic_tags(row.get('topic_tags')),
                'year_source': self._year_source(row.get('year_source')),
            })

        if len(courses_by_id) > 1:
            course_names = ', '.join(sorted(course.name for course in courses_by_id.values()))
            raise CommandError(
                'This codebase stores course filtering on ExamPaper.course, so one exit exam '
                f'paper can import one course at a time. Found multiple courses: {course_names}.'
            )

        course = next(iter(courses_by_id.values()), exam_paper.course)
        return prepared, skipped, errors, course

    def _create_questions(self, exam_paper, rows):
        created = 0
        for row in rows:
            Question.objects.create(
                exam_paper=exam_paper,
                chapter=None,
                is_active=True,
                **row,
            )
            created += 1
        return created

    def _resolve_course(self, label):
        value = self._clean(label)
        return Course.objects.filter(
            Q(name__iexact=value) | Q(code__iexact=value),
            is_active=True,
        ).first()

    def _iter_dict_rows(self, worksheet):
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return

        keys = [self._clean(value).lower() for value in header]
        for row_num, values in enumerate(rows, start=2):
            if all(self._clean(value) == '' for value in values):
                continue
            yield row_num, {
                key: values[index] if index < len(values) else None
                for index, key in enumerate(keys)
                if key
            }

    def _headers(self, worksheet):
        first_row = next(worksheet.iter_rows(max_row=1, values_only=True), None)
        if not first_row:
            return set()
        return {self._clean(value).lower() for value in first_row if self._clean(value)}

    def _is_notes_row(self, row):
        return (
            self._clean(row.get('course')).lower().startswith('required')
            or self._clean(row.get('question')).lower() == 'full question text'
            or self._clean(row.get('question_text')).lower() == 'full question text'
        )

    @staticmethod
    def _topic_tags(raw_tags):
        deduped = []
        seen = set()
        for tag in str(raw_tags or '').split(','):
            tag = tag.strip()
            if not tag or tag in seen:
                continue
            seen.add(tag)
            deduped.append(tag)
        return deduped

    @staticmethod
    def _clean(value):
        return str(value).strip() if value is not None else ''

    @staticmethod
    def _year_source(value):
        value = str(value).strip() if value is not None else ''
        field = Question._meta.get_field('year_source')
        if not value or len(value) > field.max_length:
            return ''
        return value

    def _write_errors(self, errors):
        if not errors:
            return
        self.stdout.write(self.style.WARNING('Errors:'))
        for error in errors[:20]:
            self.stdout.write(f'  {error}')
        if len(errors) > 20:
            self.stdout.write(f'  ... and {len(errors) - 20} more')
